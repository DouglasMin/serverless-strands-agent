import base64
import io
import json
import re
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from strands import tool


def _parse_cell_value(raw: str) -> Any:
    """Parse raw string into int, float, formula, or clean string."""
    s = raw.strip()
    if not s:
        return ""
    if s.startswith("="):
        return s

    # Remove currency symbol or comma
    clean = re.sub(r"^[$\€\£\₩]\s*", "", s).replace(",", "").strip()

    # Try percentage e.g. "85.5%"
    if clean.endswith("%"):
        try:
            return float(clean[:-1]) / 100.0
        except ValueError:
            pass

    # Try integer
    try:
        if re.match(r"^-?\d+$", clean):
            return int(clean)
    except ValueError:
        pass

    # Try float
    try:
        if re.match(r"^-?\d+\.\d+$", clean):
            return float(clean)
    except ValueError:
        pass

    return s


def _parse_markdown_to_sheets(md_text: str) -> list[dict[str, Any]]:
    """Parse Markdown tables and sheet headers into structured Excel sheet definitions."""
    sections = re.split(r"\n(?=##?\s+)", md_text)
    sheets: list[dict[str, Any]] = []

    for sec in sections:
        sec = sec.strip()
        if not sec:
            continue

        lines = sec.split("\n")
        sheet_name = "Data"
        title_text = ""
        first_line = lines[0].strip()

        if first_line.startswith("# ") or first_line.startswith("## "):
            sheet_name = re.sub(r"^#+\s*", "", first_line).strip()[:31]
            title_text = sheet_name

        table_lines = [l.strip() for l in lines if l.strip().startswith("|") and "|" in l.strip()[1:]]
        if len(table_lines) < 2:
            continue

        headers = [c.strip() for c in table_lines[0].split("|")[1:-1]]
        data_start = 1
        if "---" in table_lines[1] or ":---" in table_lines[1]:
            data_start = 2

        rows: list[list[Any]] = []
        summary_row: list[Any] | None = None

        for tline in table_lines[data_start:]:
            cells = [c.strip() for c in tline.split("|")[1:-1]]
            parsed_cells = [_parse_cell_value(c) for c in cells]

            # Check if this is a summary/total row
            first_cell_str = str(parsed_cells[0]).lower().strip() if parsed_cells else ""
            if first_cell_str in ["total", "totals", "sum", "average", "avg", "overall", "summary"]:
                summary_row = parsed_cells
            else:
                rows.append(parsed_cells)

        if headers and (rows or summary_row):
            sheets.append({
                "name": sheet_name,
                "title": title_text,
                "headers": headers,
                "rows": rows,
                **({"summary_row": summary_row} if summary_row else {}),
            })

    # Fallback: if no markdown headers were found but a table is present
    if not sheets:
        table_lines = [l.strip() for l in md_text.split("\n") if l.strip().startswith("|") and "|" in l.strip()[1:]]
        if len(table_lines) >= 2:
            headers = [c.strip() for c in table_lines[0].split("|")[1:-1]]
            data_start = 1
            if "---" in table_lines[1] or ":---" in table_lines[1]:
                data_start = 2
            rows = []
            summary_row = None
            for tline in table_lines[data_start:]:
                cells = [c.strip() for c in tline.split("|")[1:-1]]
                parsed_cells = [_parse_cell_value(c) for c in cells]
                first_cell_str = str(parsed_cells[0]).lower().strip() if parsed_cells else ""
                if first_cell_str in ["total", "totals", "sum", "average", "avg", "overall"]:
                    summary_row = parsed_cells
                else:
                    rows.append(parsed_cells)
            sheets.append({
                "name": "Summary",
                "title": "Data Breakdown",
                "headers": headers,
                "rows": rows,
                **({"summary_row": summary_row} if summary_row else {}),
            })

    return sheets


def _apply_table_styling(ws: Any, start_row: int, end_row: int, max_col: int):
    header_fill = PatternFill(
        start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    alt_fill = PatternFill(
        start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    summary_border = Border(
        top=Side(style="thin", color="94A3B8"),
        bottom=Side(style="double", color="1E293B"),
    )
    summary_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    summary_fill = PatternFill(
        start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"
    )

    # Style Header
    for col in range(1, max_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Style Rows
    for row in range(start_row + 1, end_row + 1):
        is_alt = (row % 2) == 0
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if cell.value is not None:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    # Format as integer or float currency if large
                    if isinstance(cell.value, float):
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "#,##0"
                else:
                    val_str = str(cell.value).strip()
                    if val_str.startswith("="):
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = "#,##0.00"
            if is_alt and row != end_row:
                cell.fill = alt_fill

    # Auto-fit Column Widths
    for col in range(1, max_col + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(start_row, end_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


@tool
def create_excel_spreadsheet(
    filename: str,
    sheets_json: str | None = None,
    content: str | None = None,
) -> str:
    """Create a styled Microsoft Excel (.xlsx) spreadsheet with multiple sheets, formatted tables, and formulas.

    Supports either structured sheets_json OR raw Markdown tables string.

    filename: Name of the file, e.g. "Q3_Financial_Model.xlsx"
    content: Raw Markdown tables text (sheets separated by '## Sheet Name')
    sheets_json: Optional JSON string representing sheets. Example:
    [
      {
        "name": "Revenue",
        "title": "Quarterly Revenue Breakdown (USD)",
        "headers": ["Quarter", "Product A", "Product B", "Total Revenue"],
        "rows": [
          ["Q1 2026", 120000, 85000, "=B4+C4"],
          ["Q2 2026", 145000, 92000, "=B5+C5"],
          ["Q3 2026", 160000, 105000, "=B6+C6"],
          ["Q4 2026", 190000, 120000, "=B7+C7"]
        ],
        "summary_row": ["Total", "=SUM(B4:B7)", "=SUM(C4:C7)", "=SUM(D4:D7)"]
      }
    ]
    """
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    sheets_data: list[dict[str, Any]] = []

    # 1. Try parsing direct Markdown content if provided
    if content and isinstance(content, str) and content.strip():
        sheets_data = _parse_markdown_to_sheets(content)
    # 2. Try parsing sheets_json
    elif sheets_json and isinstance(sheets_json, str) and sheets_json.strip():
        try:
            parsed = json.loads(sheets_json)
            if isinstance(parsed, list):
                sheets_data = parsed
            elif isinstance(parsed, dict):
                sheets_data = [parsed]
        except Exception:
            # Fallback: treat sheets_json as raw markdown tables
            sheets_data = _parse_markdown_to_sheets(sheets_json)

    if not sheets_data:
        sheets_data = [
            {
                "name": "Sheet1",
                "title": "Data Overview",
                "headers": ["Item", "Value"],
                "rows": [["Sample A", 100], ["Sample B", 200]],
            }
        ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    total_sheets = 0
    total_rows = 0

    title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")

    for sheet_idx, s in enumerate(sheets_data):
        sheet_name = s.get("name", f"Sheet{sheet_idx + 1}")[:31]
        ws = wb.create_sheet(title=sheet_name)
        total_sheets += 1

        curr_row = 1
        title_text = s.get("title")
        if title_text:
            title_cell = ws.cell(row=curr_row, column=1, value=title_text)
            title_cell.font = title_font
            curr_row += 2

        headers = s.get("headers", [])
        if not headers:
            continue

        header_row_idx = curr_row
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=header_row_idx, column=col_idx, value=str(h))
        curr_row += 1

        rows = s.get("rows", [])
        for r in rows:
            for col_idx, val in enumerate(r, 1):
                if col_idx <= len(headers):
                    ws.cell(row=curr_row, column=col_idx, value=val)
            curr_row += 1
            total_rows += 1

        summary = s.get("summary_row")
        if summary:
            for col_idx, val in enumerate(summary, 1):
                if col_idx <= len(headers):
                    cell = ws.cell(row=curr_row, column=col_idx, value=val)
                    cell.font = Font(
                        name="Calibri", size=11, bold=True, color="0F172A"
                    )
            curr_row += 1
            total_rows += 1

        _apply_table_styling(
            ws,
            start_row=header_row_idx,
            end_row=curr_row - 1,
            max_col=len(headers),
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    size_bytes = len(excel_bytes)
    base64_data = base64.b64encode(excel_bytes).decode("utf-8")
    data_uri = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{base64_data}"

    doc_event = {
        "filename": filename,
        "file_type": "excel",
        "size_bytes": size_bytes,
        "data_uri": data_uri,
        "summary": f"{total_sheets} sheet(s), {total_rows} row(s)",
    }

    from office_tools import document_queue

    document_queue.put_nowait(doc_event)

    return json.dumps(
        {
            "status": "success",
            "filename": filename,
            "file_type": "excel",
            "size_bytes": size_bytes,
            "sheets_count": total_sheets,
            "download_ready": True,
            "summary": f"Successfully created Excel spreadsheet '{filename}' with {total_sheets} sheet(s).",
        },
        indent=2,
    )
